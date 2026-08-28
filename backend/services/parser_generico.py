"""Parser genérico Excel/CSV guiado por PlantillaCarga.

Esquema JSON de configuración (guardado en PlantillaCarga.configuracion):
{
  "header_row": 0,
  "columnas": {
    "codigo_colaborador": "CODIGO" | {"names": ["CODIGO","COD"]},
    "nombre_colaborador": "NOMBRE",
    "fecha": "FECHA",
    "labor": "LABOR",            (opcional si 'labor_fija')
    "tallos": "TALLOS",          (opcional)
    "ramos": "RAMOS",            (opcional)
    "horas_ordinarias": "HS_ORD",
    "horas_extra_ordinarias": "HE_ORD",
    "horas_dominicales": "HS_DOM",
    "unidades_tarea": "UNID_TAREA",
    "horas_tarea": "HS_TAREA",
    "semana": "SEMANA"           (opcional; si falta se deriva de fecha)
  },
  "labor_fija": "CORTE",
  "fecha_formato": null   // si null: parseo automático
}
"""

from __future__ import annotations

import io
import json
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from services.utils_semana import (
    normalizar_codigo_semana,
    semana_desde_fecha,
    tallos_a_ramos,
)


CAMPOS_NUMERICOS = (
    "tallos", "ramos",
    "horas_ordinarias", "horas_extra_ordinarias", "horas_dominicales",
    "unidades_tarea", "horas_tarea",
)


def _formato_mes_primero(fmt: str | None) -> bool:
    """Detecta si el number_format de Excel pone el mes antes del día (locale US)."""
    fl = (fmt or "").lower()
    if "d" not in fl or "m" not in fl:
        return False
    # Ignorar minutos: en Excel 'mm' después de ':' o 'hh' es minuto, no mes.
    # Heurística simple: tomar el primer 'd' y primer 'm' en la parte de fecha.
    return fl.index("m") < fl.index("d")


def _leer_excel_openpyxl(contenido: bytes, header_row: int, dia_primero: bool) -> pd.DataFrame:
    """Lee xlsx con openpyxl preservando formatos. Para celdas de fecha ambiguas
    (día ≤ 12) asume que Excel las interpretó con locale US y las invierte para
    recuperar la intención día-primero del usuario colombiano.
    """
    wb = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))
    if not rows or header_row >= len(rows):
        return pd.DataFrame()
    header = [str(c.value).strip() if c.value is not None else "" for c in rows[header_row]]
    data = []
    for row in rows[header_row + 1:]:
        record = {}
        for idx, cell in enumerate(row):
            if idx >= len(header) or not header[idx]:
                continue
            val = cell.value
            # Serial de Excel (celda formateada como 'General' pero con fecha)
            if isinstance(val, (int, float)) and not isinstance(val, bool) and 40000 < val < 60000:
                fmt = (cell.number_format or "").lower()
                if "d" in fmt or "m" in fmt or "y" in fmt or fmt in ("general", ""):
                    try:
                        val = from_excel(val)
                    except Exception:
                        pass
            if isinstance(val, datetime):
                # Excel ya resolvió la fecha al abrir el archivo según el locale
                # de quien lo guardó. Confiamos en el datetime tal cual.
                val = val.strftime("%d/%m/%Y")
            record[header[idx]] = val
        if any(v not in (None, "", " ") for v in record.values()):
            data.append(record)
    return pd.DataFrame(data)


def _leer_archivo(contenido: bytes, nombre: str, header_row: int, dia_primero: bool = True) -> pd.DataFrame:
    nombre_lower = nombre.lower()
    if nombre_lower.endswith(".csv"):
        return pd.read_csv(io.BytesIO(contenido), header=header_row, dtype=str, keep_default_na=False)
    return _leer_excel_openpyxl(contenido, header_row, dia_primero)


def _resolver_nombre_col(df: pd.DataFrame, spec: Any) -> str | None:
    """Encuentra la columna real a partir del spec (string o {'names': [...]})."""
    if spec is None:
        return None
    candidatos = [spec] if isinstance(spec, str) else spec.get("names", [])
    # matching case-insensitive y tolerante a espacios
    cols_norm = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidatos:
        k = str(cand).strip().lower()
        if k in cols_norm:
            return cols_norm[k]
    return None


def _parse_float(v) -> float:
    if v is None or v == "" or (isinstance(v, float) and pd.isna(v)):
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _parse_fecha(v, fmt: str | None) -> date | None:
    if v is None or v == "" or (isinstance(v, float) and pd.isna(v)):
        return None
    # Objetos datetime/date (Excel puede devolverlos directos)
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime().date()
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() in ("nan", "nat", "none"):
        return None
    if fmt:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            return None
    # Prioridad: DD/MM/YYYY (formato colombiano), luego variantes
    for probar in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, probar).date()
        except ValueError:
            continue
    # Fallback con pandas (dayfirst=True fuerza interpretación DD/MM)
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.to_pydatetime().date()
    except Exception:
        return None


def parsear(contenido: bytes, nombre_archivo: str, plantilla_config: str | dict,
            tallos_por_ramo: int = 1) -> dict:
    """Parsea un archivo y devuelve {'registros': [...], 'errores': [...]}.

    Cada registro es un dict compatible con models.RegistroDiario.
    Convierte tallos→ramos automáticamente si hay 'tallos' y no hay 'ramos'.
    """
    cfg = json.loads(plantilla_config) if isinstance(plantilla_config, str) else plantilla_config
    header_row = cfg.get("header_row", 0)
    columnas_cfg = cfg.get("columnas", {})
    labor_fija = cfg.get("labor_fija")
    fecha_fmt = cfg.get("fecha_formato")
    dia_primero = cfg.get("fecha_dia_primero", True)

    try:
        df = _leer_archivo(contenido, nombre_archivo, header_row, dia_primero)
    except Exception as e:
        return {"registros": [], "errores": [{"fila": 0, "error": f"No se pudo leer archivo: {e}"}]}

    mapeo = {campo: _resolver_nombre_col(df, spec) for campo, spec in columnas_cfg.items()}

    # Campos obligatorios
    faltantes = [
        campo for campo in ("codigo_colaborador", "nombre_colaborador", "fecha")
        if not mapeo.get(campo)
    ]
    if faltantes:
        return {"registros": [], "errores": [
            {"fila": 0, "error": f"Columnas obligatorias no encontradas: {', '.join(faltantes)}"}
        ]}
    if not labor_fija and not mapeo.get("labor"):
        return {"registros": [], "errores": [
            {"fila": 0, "error": "Se requiere columna 'labor' o 'labor_fija' en la plantilla"}
        ]}

    registros, errores = [], []
    for idx, fila in df.iterrows():
        nro_fila = int(idx) + header_row + 2  # fila humana en Excel
        try:
            codigo_raw = fila[mapeo["codigo_colaborador"]]
            if codigo_raw is None or str(codigo_raw).strip() == "":
                continue  # fila vacía, saltar silenciosamente
            codigo = int(float(str(codigo_raw).strip()))
            nombre = str(fila[mapeo["nombre_colaborador"]]).strip()
            fecha = _parse_fecha(fila[mapeo["fecha"]], fecha_fmt)
            if not fecha:
                errores.append({"fila": nro_fila, "error": "Fecha inválida"})
                continue

            labor = labor_fija or str(fila[mapeo["labor"]]).strip()

            semana_cod = None
            if mapeo.get("semana") and fila.get(mapeo["semana"]):
                try:
                    semana_cod = normalizar_codigo_semana(fila[mapeo["semana"]])
                except ValueError:
                    pass
            if not semana_cod:
                semana_cod = semana_desde_fecha(fecha)

            vals = {c: _parse_float(fila[mapeo[c]]) if mapeo.get(c) else 0.0
                    for c in CAMPOS_NUMERICOS}

            # Si vinieron tallos pero no ramos, convertir
            if vals["tallos"] and not vals["ramos"]:
                vals["ramos"] = tallos_a_ramos(vals["tallos"], tallos_por_ramo)

            registros.append({
                "codigo_colaborador": codigo,
                "nombre_colaborador": nombre,
                "fecha": fecha,
                "semana": semana_cod,
                "labor": labor,
                **vals,
            })
        except Exception as e:
            errores.append({"fila": nro_fila, "error": str(e)})

    return {"registros": registros, "errores": errores}
